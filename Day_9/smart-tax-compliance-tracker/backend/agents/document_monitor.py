import os
import glob
from PyPDF2 import PdfReader
from docx import Document
import csv
import json
import openpyxl
from backend.config import Config
from backend.rag.vector_store import VectorStore

class DocumentMonitorAgent:
    def __init__(self):
        self.vector_store = VectorStore()
        self.processed_files = set()
        self._load_processed_files()

    def _load_processed_files(self):
        try:
            with open("backend/data/processed_files.txt", "r") as f:
                self.processed_files = set(f.read().splitlines())
        except FileNotFoundError:
            pass

    def _save_processed_files(self):
        with open("backend/data/processed_files.txt", "w") as f:
            f.write("\n".join(self.processed_files))

    def _extract_text(self, file_path):
        text = ""
        if file_path.endswith('.pdf'):
            reader = PdfReader(file_path)
            text = "\n".join([page.extract_text() for page in reader.pages])
        elif file_path.endswith('.docx'):
            doc = Document(file_path)
            text = "\n".join([para.text for para in doc.paragraphs])
        elif file_path.endswith('.txt'):
            with open(file_path, 'r') as f:
                text = f.read()
        elif file_path.endswith('.csv'):
            with open(file_path, 'r') as f:
                reader = csv.reader(f)
                text = "\n".join([",".join(row) for row in reader])
        elif file_path.endswith('.json'):
            with open(file_path, 'r') as f:
                data = json.load(f)
                text = json.dumps(data)
        elif file_path.endswith(('.xlsx', '.xls')):
            wb = openpyxl.load_workbook(file_path)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                text += f"Sheet: {sheet}\n"
                for row in ws.iter_rows(values_only=True):
                    text += ",".join([str(cell) for cell in row]) + "\n"
        return text

    def monitor_documents(self):
        new_files = []
        paths = [
            Config.INVOICES_PATH,
            Config.TRANSACTIONS_PATH,
            Config.LEDGERS_PATH,
            Config.REGULATIONS_PATH
        ]

        for path in paths:
            for ext in ('*.pdf', '*.docx', '*.txt', '*.csv', '*.json', '*.xlsx'):
                for file in glob.glob(os.path.join(path, ext)):
                    if file not in self.processed_files:
                        new_files.append(file)
                        self.processed_files.add(file)

        if new_files:
            documents = []
            for file in new_files:
                text = self._extract_text(file)
                documents.append(f"File: {file}\nContent:\n{text}")
            
            self.vector_store.add_documents(documents)
            self._save_processed_files()
        
        return f"Processed {len(new_files)} new files"